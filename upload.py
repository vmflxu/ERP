import json
import openpyxl as op
from PyQt5.QtWidgets import QDialog, QFileDialog
from PyQt5.QtCore import QSettings
from PyQt5 import uic
from firebase_admin import firestore
from datetime import datetime

UI_upload = uic.loadUiType("ui_upload_po.ui")[0]

db = firestore.client()



#   B4: order#, E5 : 작성날짜, E6: 완료날짜, H4: 작성자, I4: 책임자
#   작업내역 항목별 좌표
#   row : 8,10,12,14,16,18,20,22
#   column : 3~8

class UploadDialog(QDialog,UI_upload):
    
    path = "poform.xlsx"
    wb = op.load_workbook(path)
    ws = wb.active
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    ws.print_options.print_area = 'B2:I28'
    # ws.print_area = 'B2:I28'
    dataset ={}
    B4 = None
    E5 = None
    E6 = None
    H4 = None
    I4 = None
    # C24 = None
    save_path = None
    savepath = None

    def __init__(self):
        super().__init__()
        self.initUI()
        
        print(op.__version__)
        self.button_Cancle.clicked.connect(self.cancle)
        self.button_Path.clicked.connect(self.setPath)
        self.button_Upload.clicked.connect(self.executeUpload)
        # self.show()
    
    def initUI(self):
        settings = QSettings('TSC','TSC ERP')
        self.setupUi(self)
        # self.check_Save.setCheckState(settings.value('엑셀파일 저장',2))
        self.label_Path.setText(settings.value('경로',""))
# settings.value('엑셀파일 저장',True)
    def setPath(self):
        self.savepath = self.label_Path.text()
        self.savepath = QFileDialog.getExistingDirectory(self,"경로를 지정하세요")        
        self.label_Path.setText(self.savepath)
        
    def executeUpload(self):
        print(self.check_Save.isChecked())
        #   체크박스에 따라서 Excel 저장
        self.savepath = self.label_Path.text()
        if self.check_Save.isChecked() == True:
            filename = "/"+str(self.dataset["START_DAY"]).rstrip('0')+str(self.dataset["NUMBER"])+".xlsx"
            self.save_path = self.savepath + filename
            self.passToExcel()
            self.wb.save(self.save_path)

        #   DB 업로드(무조건 실행)
        self.dataset = json.loads(json.dumps(self.dataset,indent=4,ensure_ascii=False))

        db.collection('PO_PJT').document().set(self.dataset)
        self.button_Upload.setDisabled(True)
        self.label_Upload.setText("업로드가 완료되었습니다.")
        self.button_Cancle.setText("닫기(ESC)")

    def passToExcel(self):        
        self.B4 = str(self.dataset["NUMBER"])
        self.E5 = self.dateEdit(self.dataset["START_DAY"])
        self.E6 = self.dateEdit(self.dataset["END_DAY"])
        self.D7 = str(self.dataset["CUSTOMER"])
        self.H7 = str(self.dataset["PJT"])
        # self.E6 = str(self.dataset["END_DAY"].rstrip('0').rstrip('_'))
        self.H4 = str(self.dataset["ORDERER"])
        self.C25 = str(self.dataset["SPECIALPOINT"])
        self.setCoodinates()
        temp = self.dataset["WORKLIST"]

        self.i = 0
        for rows in range(9,24,2):
            for columns in range(3,9,1):
                self.ws.cell(row=rows, column=columns).value = '' if str(temp[self.i][columns-3]) == '0' else str(temp[self.i][columns-3])
            self.i = self.i + 1
    
    def dateEdit(self,data):
        temp = data.rstrip('0').rstrip('_')
        year = temp[0:4]
        month = temp[4:6]
        date = temp[6:8]
        return year+'.'+month+'.'+date
    def cancle(self):
        self.close()

    def setCoodinates(self):
        # self.cell(row=4,column=2).value = self.B4
        # self.cell(row=5,column=5).value = self.E5
        # self.cell(row=6,column=5).value = self.E6
        # self.cell(row=4,column=6).value = self.H4
        # self.cell(row=4,column=7).value = self.I4
        # self.cell(row=24,column=3).value = self.C24
        self.ws['B4'] = self.B4
        self.ws['E5'] = self.E5
        self.ws['E6'] = self.E6
        self.ws['H4'] = self.H4
        self.ws['I4'] = self.I4
        self.ws['D7'] = self.D7
        self.ws['H7'] = self.H7
        self.ws['C25'] = self.C25
    
    def closeEvent(self, event):
        settings = QSettings('TSC','TSC ERP')
        # settings.setValue('엑셀파일 저장', self.check_Save.checkState())
        # settings.setValue('프린트 출력', self.check_Print.checkState())
        settings.setValue('경로',self.label_Path.text())